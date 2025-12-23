import sys
import os
import shutil
import traceback
import logging
import zipfile
import time
import pandas as pd
import pandas_gbq
import unicodedata
import math
import getpass
import openpyxl
import win32com.client as win32
import pythoncom
import pytz
from pathlib import Path
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor
from google.cloud import bigquery
from typing import Dict, List, Tuple, Optional

# ==================================================================================================
# CONSTANTES E CONFIGURAÇÕES
# ==================================================================================================

SCRIPT_NAME = "finalizadosfgts"
AREA_NAME = "BO CONSIGNADO"

TZ = pytz.timezone("America/Sao_Paulo")
ENV_EXEC_USER = os.getenv("ENV_EXEC_USER", getpass.getuser()).lower()
ENV_EXEC_MODE = os.getenv("ENV_EXEC_MODE", "MANUAL").upper()
TEST_MODE = False

# Paths - Robust Detection
HOME = Path.home()
POSSIBLE_ROOTS = [
    HOME / "C6 CTVM LTDA, BANCO C6 S.A. e C6 HOLDING S.A",
    HOME / "Meu Drive/C6 CTVM",
    HOME / "C6 CTVM",
    HOME / "SharePoint",
    HOME / "OneDrive - C6 Bank S.A",
    HOME,
]
ROOT_DRIVE = next((p for p in POSSIBLE_ROOTS if p.exists() and (p / "Mensageria e Cargas Operacionais - 11.CelulaPython").exists()), POSSIBLE_ROOTS[0])

# Specific Paths
BASE_DIR = ROOT_DRIVE / "Mensageria e Cargas Operacionais - 11.CelulaPython/graciliano"
if not BASE_DIR.exists():
    BASE_DIR = ROOT_DRIVE / "graciliano"

AUTOMACOES_DIR = BASE_DIR / "automacoes"
LOG_DIR = AUTOMACOES_DIR / AREA_NAME / "LOGS" / SCRIPT_NAME / datetime.now(TZ).strftime("%Y-%m-%d")
TEMP_DIR = Path(os.getenv('TEMP', Path.home())) / "C6_RPA_EXEC" / SCRIPT_NAME

# Business Paths
INPUT_PATH = ROOT_DRIVE / "BKO FINANCEIRO - ENVIO DE CONTRATOS PARA BAIXA" / "ENVIADOS PARA BAIXA - GERAL 2025.xlsx"

# BigQuery
PROJECT_ID = "datalab-pagamentos"
DATASET_ID = "ADMINISTRACAO_CELULA_PYTHON"

REGISTRO_AUTOMACOES_TABLE = f"{PROJECT_ID}.ADMINISTRACAO_CELULA_PYTHON.registro_automacoes"
AUTOMACOES_EXEC_TABLE = f"{PROJECT_ID}.ADMINISTRACAO_CELULA_PYTHON.automacoes_exec"

BQ_DATASET_NEGOCIO = "conciliacoes_monitoracao"
BQ_TABLE_NEGOCIO = "enviados_para_baixa"
TABELA_FINAL = f"{PROJECT_ID}.{BQ_DATASET_NEGOCIO}.{BQ_TABLE_NEGOCIO}"

COL_ORIG = ["CHAVE (OPERAÇÃO&VALOR)","CHAVE (CPF&OPERAÇÃO&VALOR)","CPF","OPERAÇÃO","DATA PREVISTA REPASSE"," VALOR CEDIDO/ALIENADO ORIGINAL EDITADO ","DATA EFETIVA DE PAGAMENTO","SITUAÇÃO","AÇÃO","DATA SOLICITAÇÃO BAIXA","ARQUIVO"]
COL_DEST = ["CHAVE_OPERACAO_VALOR","CHAVE_CPF_OPERACAO_VALOR","CPF","OPERACAO","DATA_PREVISTA_REPASSE","VALOR_CEDIDO_ALIENADO_ORIGINAL_EDITADO","DATA_EFETIVA_DE_PAGAMENTO","SITUACAO","ACAO","DATA_SOLICITACAO_BAIXA","ARQUIVO"]
MAPA_DEST = dict(zip(COL_ORIG, COL_DEST))

# ==================================================================================================
# FUNÇÕES DE SUPORTE
# ==================================================================================================

def setup_logger():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    
    log_filename = f"{SCRIPT_NAME}_{datetime.now(TZ).strftime('%H%M%S')}.log"
    log_path = LOG_DIR / log_filename
    
    logger = logging.getLogger(SCRIPT_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers = []
    
    formatter = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    
    file_handler = logging.FileHandler(log_path, encoding='utf-8')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)
    
    return logger, log_path

def load_config(logger):
    if TEST_MODE:
        return {"emails_principal": ["carlos.lsilva@c6bank.com"], "emails_cc": [], "move_file": False, "is_active": True}
    try:
        query = f"""
            SELECT emails_principal, emails_cc, move_file, is_active
            FROM `{REGISTRO_AUTOMACOES_TABLE}`
            WHERE lower(TRIM(script_name)) = lower('{SCRIPT_NAME}')
            ORDER BY created_at DESC LIMIT 1
        """
        try:
            df = pandas_gbq.read_gbq(query, project_id=PROJECT_ID, use_bqstorage_api=False)
        except Exception:
             # Fallback
            query = query.replace(f"lower('{SCRIPT_NAME}')", f"lower('{AREA_NAME.lower()}')")
            df = pandas_gbq.read_gbq(query, project_id=PROJECT_ID, use_bqstorage_api=False)

        if not df.empty:
            row = df.iloc[0]
            val_move = row.get("move_file", False)
            if isinstance(val_move, str): val_move = val_move.lower() in ('true', '1')
            else: val_move = bool(val_move)
            
            val_active = row.get("is_active", "true")
            if isinstance(val_active, str): val_active = val_active.lower() in ('true', '1', 'ativo')
            else: val_active = bool(val_active)

            return {
                "emails_principal": [e.strip() for e in row.get("emails_principal", "").split(";") if "@" in e],
                "emails_cc": [e.strip() for e in row.get("emails_cc", "").split(";") if "@" in e],
                "move_file": val_move,
                "is_active": val_active
            }
    except Exception as e:
        logger.warning(f"Erro config BQ: {e}")
    return {"emails_principal": ["carlos.lsilva@c6bank.com"], "emails_cc": [], "move_file": False, "is_active": True}

def record_metrics(logger, start_time, end_time, status, error_msg=""):
    if TEST_MODE: return
    try:
        duration = (end_time - start_time).total_seconds()
        metrics = {
            "script_name": SCRIPT_NAME,
            "area_name": AREA_NAME,
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "duration_seconds": duration,
            "status": status,
            "usuario": ENV_EXEC_USER,
            "modo_exec": ENV_EXEC_MODE,
        }
        pandas_gbq.to_gbq(pd.DataFrame([metrics]), AUTOMACOES_EXEC_TABLE, project_id=PROJECT_ID, if_exists="append", use_bqstorage_api=False)
    except Exception as e:
        logger.error(f"Erro metrics: {e}")

def send_email_outlook(logger, subject, body, to_list, cc_list=None, attachments=None):
    if not to_list or TEST_MODE: return
    try:
        pythoncom.CoInitialize()
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        if body: mail.HTMLBody = body
        mail.To = ";".join(to_list)
        if cc_list: mail.CC = ";".join(cc_list)
        if attachments:
            for att in attachments:
                if Path(att).exists(): mail.Attachments.Add(str(att))
        mail.Send()
    except Exception as e:
        logger.error(f"Erro email: {e}")
    finally:
         try: pythoncom.CoUninitialize()
         except: pass

def smart_zip_logs(output_files: list) -> str:
    zip_filename = f"{SCRIPT_NAME}_{datetime.now(TZ).strftime('%H%M%S')}.zip"
    zip_path = TEMP_DIR / zip_filename
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        current_logs = list(LOG_DIR.glob(f"{SCRIPT_NAME}_*.log"))
        if current_logs:
            latest = max(current_logs, key=os.path.getctime)
            zf.write(latest, arcname=latest.name)
            
        for f in output_files:
            if Path(f).exists():
                zf.write(f, arcname=Path(f).name)
                
    return str(zip_path)

# ==================================================================================================
# LÓGICA DE NEGÓCIO
# ==================================================================================================

def remover_acentos(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")

def normalizar_rotulo(s: str) -> str:
    t = str(s or "").lower().strip()
    t = remover_acentos(t)
    for ch in ["-", "/", "\\", ",", ";", ".", ":", "(", ")", "[", "]"]:
        t = t.replace(ch, " ")
    t = t.replace("&", " E ")
    while "  " in t:
        t = t.replace("  ", " ")
    for stop in (" DE ", " DO ", " DA "):
        t = t.replace(stop, " ")
    return " ".join(t.split())

def adivinhar_origem(nk: str) -> Optional[str]:
    nk = str(nk or "").lower().strip()
    if "CHAVE" in nk and "CPF" in nk and "OPERACAO" in nk and "VALOR" in nk:
        return "CHAVE (CPF&OPERAÇÃO&VALOR)"
    if "CHAVE" in nk and "OPERACAO" in nk and "VALOR" in nk:
        return "CHAVE (OPERAÇÃO&VALOR)"
    if nk.startswith("CPF"):
        return "CPF"
    if "OPERACAO" in nk and "CHAVE" not in nk:
        return "OPERAÇÃO"
    if "DATA" in nk and "PREVISTA" in nk and "REPASSE" in nk:
        return "DATA PREVISTA REPASSE"
    if "VALOR" in nk and ("CEDIDO" in nk or "ALIENADO" in nk) and "EDITADO" in nk:
        return " VALOR CEDIDO/ALIENADO ORIGINAL EDITADO "
    if "DATA" in nk and "EFETIVA" in nk and "PAGAMENTO" in nk:
        return "DATA EFETIVA DE PAGAMENTO"
    if "SITUACAO" in nk:
        return "SITUAÇÃO"
    if "ACAO" in nk:
        return "AÇÃO"
    if "DATA" in nk and "SOLICITACAO" in nk and "BAIXA" in nk:
        return "DATA SOLICITAÇÃO BAIXA"
    if "ARQUIVO" in nk:
        return "ARQUIVO"
    return None

def formatar_valor_como_texto(valor, fmt: str) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) or isinstance(valor, int):
        fmt_up = (fmt or "").lower()
        if "," in fmt_up or "[$" in fmt_up or "0,00" in fmt_up or "_-" in fmt_up:
            if isinstance(valor, float) and not math.isnan(valor):
                s2 = f"{valor:,.2f}"
            else:
                s2 = f"{float(valor):,.0f}"
            s2 = s2.replace(",", "X").replace(".", ",").replace("X", ".")
            return s2
        return str(valor)
    return str(valor).strip()

def detecting_cabecalho(linhas: List[List[str]], max_scan: int = 80) -> Tuple[Optional[int], Dict[int, str]]:
    melhor_score = 0
    melhor_idx = None
    melhor_map: Dict[int, str] = {}
    linhas_scan = min(max_scan, len(linhas))
    
    assinaturas_locais = {c: normalizar_rotulo(c) for c in COL_ORIG}
    
    for r in range(linhas_scan):
        linha = [str(x or "").strip() for x in (linhas[r] if r < len(linhas) else [])]
        norm = [normalizar_rotulo(x) for x in linha]
        pos2orig: Dict[int, str] = {}
        usados = set()
        score = 0
        for i, nk in enumerate(norm):
            if not nk:
                continue
            hit = None
            for orig, sig in assinaturas_locais.items():
                if orig in usados:
                    continue
                if nk == sig:
                    hit = orig
                    break
            if not hit:
                guess = adivinhar_origem(nk)
                if guess and guess not in usados:
                    hit = guess
            if hit:
                pos2orig[i] = hit
                usados.add(hit)
                score += 1
        if score > melhor_score:
            melhor_score = score
            melhor_idx = r
            melhor_map = pos2orig
        if score == len(COL_ORIG):
            break
    return melhor_idx, melhor_map

def ler_planilha_sem_ui(caminho: Path, logger) -> Dict[str, List[List[str]]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl não instalado.")
    wb = openpyxl.load_workbook(filename=str(caminho), data_only=True, read_only=True)
    dados: Dict[str, List[List[str]]] = {}
    for ws in wb.worksheets:
        linhas = ws.max_row or 0
        colunas = min(ws.max_column or 0, 11)
        bloco: List[List[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=linhas, min_col=1, max_col=colunas):
            linha_txt: List[str] = []
            for cell in row:
                linha_txt.append(formatar_valor_como_texto(cell.value, cell.number_format))
            bloco.append(linha_txt)
        dados[ws.title] = bloco
    wb.close()
    logger.info(f"Arquivo lido em memória | abas={len(dados)}")
    return dados

def escolher_aba_e_header(caminho: Path, logger) -> Tuple[List[List[str]], str, int, Dict[int, str]]:
    bruto = ler_planilha_sem_ui(caminho, logger)
    melhor = (0, None, None, None)
    for nome, linhas in bruto.items():
        linhas = [l for l in linhas if any(str(x).strip() for x in l)]
        hidx, pmap = detecting_cabecalho(linhas)
        if hidx is None or not pmap:
            continue
        score = len(pmap)
        if score > melhor[0]:
            melhor = (score, nome, hidx, pmap)
    if melhor[1] is None:
        raise ValueError("Cabeçalho não localizado em nenhuma aba")
    logger.info(f"Aba escolhida: '{melhor[1]}' | Header Row: {melhor[2]}")
    return bruto[melhor[1]], melhor[1], int(melhor[2]), dict(melhor[3])

def tratar_dataframe(caminho_xlsx: Path, logger) -> pd.DataFrame:
    linhas, aba, header_idx, pos2orig = escolher_aba_e_header(caminho_xlsx, logger)
    posicoes_ordenadas = sorted(pos2orig.keys())
    nomes_origem = [pos2orig[i] for i in posicoes_ordenadas]
    
    dados: List[List[str]] = []
    for r in range(header_idx + 1, len(linhas)):
        linha = linhas[r]
        rec = [("" if i >= len(linha) else str(linha[i]).strip()) for i in posicoes_ordenadas]
        dados.append(rec)
    df = pd.DataFrame(dados, columns=nomes_origem)
    
    for nome in COL_ORIG:
        if nome not in df.columns:
            df[nome] = ""
    df = df[COL_ORIG].astype(str)
    
    # Recorte C..K (Lógica original)
    cols_ck = COL_ORIG[2:]
    mask_valida = df[cols_ck].apply(lambda r: "".join([str(x or "") for x in r]).strip() != "", axis=1)
    if not mask_valida.any():
        logger.warning("Nenhuma linha válida nas colunas C..K")
        df = df.iloc[0:0]
    else:
        mask_np = mask_valida.to_numpy()
        ultimo_pos = int((mask_np.nonzero()[0]).max())
        df = df.iloc[:ultimo_pos + 1]
        df = df.iloc[mask_np[:ultimo_pos + 1]]
        
    logger.info(f"Linhas após recorte: {len(df)}")
    df.rename(columns={orig: MAPA_DEST[orig] for orig in COL_ORIG}, inplace=True)
    return df.reset_index(drop=True)

def main():
    start_time = datetime.now(TZ)
    logger, log_path = setup_logger()
    logger.info(f"Inicio {SCRIPT_NAME}")
    
    status = "SUCESSO"
    error_msg = ""
    output_files = [] 
    novas_linhas = 0
    
    try:
        config = load_config(logger)
        
        if not INPUT_PATH.exists():
            status = "ERRO_ARQUIVO"
            logger.error(f"Excel não encontrado: {INPUT_PATH}")
        else:
            logger.info(f"Processando: {INPUT_PATH}")
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(tratar_dataframe, INPUT_PATH, logger)
                df_processado = future.result()
            
            if df_processado.empty:
                status = "NO_DATA"
                logger.info("DataFrame vazio processado.")
            else:
                # BigQuery Upload with Staging+Merge
                client_bq = bigquery.Client(project=PROJECT_ID)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                tabela_stg = f"{PROJECT_ID}.{BQ_DATASET_NEGOCIO}.{BQ_TABLE_NEGOCIO}_stg_{timestamp}"
                
                logger.info(f"Staging: {tabela_stg}")
                pandas_gbq.to_gbq(
                    df_processado, 
                    destination_table=f"{BQ_DATASET_NEGOCIO}.{BQ_TABLE_NEGOCIO}_stg_{timestamp}",
                    project_id=PROJECT_ID,
                    if_exists='replace',
                    use_bqstorage_api=False,
                    table_schema=[{'name': col, 'type': 'STRING'} for col in df_processado.columns]
                )
                
                colunas_sql = ", ".join([f"`{c}`" for c in df_processado.columns])
                condicoes = " AND ".join([f"T.`{c}` = S.`{c}`" for c in df_processado.columns])
                
                sql_merge = f"""
                MERGE `{TABELA_FINAL}` T
                USING `{tabela_stg}` S
                ON {condicoes}
                WHEN NOT MATCHED THEN
                  INSERT ({colunas_sql})
                  VALUES ({colunas_sql})
                """
                
                logger.info("Executando MERGE BigQuery...")
                job = client_bq.query(sql_merge)
                job.result()
                
                novas_linhas = job.num_dml_affected_rows if job.num_dml_affected_rows else 0
                logger.info(f"MERGE Concluído. Novas Linhas: {novas_linhas}")
                
                # Cleanup Staging
                client_bq.delete_table(tabela_stg, not_found_ok=True)
                
                if novas_linhas == 0:
                    status = "NO_DATA"
                
    except Exception as e:
        status = "ERRO"
        error_msg = str(e)
        logger.error(f"Fatal: {traceback.format_exc()}")
        
    end_time = datetime.now(TZ)
    zip_path = smart_zip_logs(output_files)
    
    body = f"""
    <html><body>
    <h2>Execução {SCRIPT_NAME}</h2>
    <p>Status: {status}</p>
    <p>Linhas Inseridas: {novas_linhas}</p>
    </body></html>
    """
    send_email_outlook(logger, f"{SCRIPT_NAME} - {status}", body, config["emails_principal"], config["emails_cc"], [zip_path])
    record_metrics(logger, start_time, end_time, status, error_msg)

if __name__ == "__main__":
    if "--test" in sys.argv: TEST_MODE = True
    main()
